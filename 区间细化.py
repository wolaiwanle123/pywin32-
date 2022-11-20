#此函数功能为实现把路面类型自动匹配到基础信息表中,功能尚未实现
#paixu函数实现把列表去重并按从小到大进行排序
def paixu(x):
    x_sorted = sorted(x)
    x_sorted_kind_2 = []
    for i in x_sorted:
        if (i not in x_sorted_kind_2):
            x_sorted_kind_2.append(i)
    return x_sorted_kind_2

import openpyxl
wb = openpyxl.load_workbook(r"D:\03 python代码\03 区间细化\区间信息导入模版-庆阳市 - 1.xlsx")  #打开区间表
sheet1 = wb['区间表']
sheet2 = wb['路面类型表']
sheet3 = wb['新区间表']
sheet4 = wb['临时表']
row1 = sheet1.max_row
row2 = sheet2.max_row
k = 2
for i in range(2,row1+1):
    a2 = sheet1['A{}'.format(i)].value
    b2 = sheet1['B{}'.format(i)].value
    c2 = sheet1['C{}'.format(i)].value


    row4_1 = sheet4.max_row
    for j in range(2,row2+1):
        if a2 == sheet2['A{}'.format(j)].value:
            sheet4['A{}'.format(k)] =sheet2['A{}'.format(j)].value
            sheet4['B{}'.format(k)] = sheet2['B{}'.format(j)].value
            sheet4['C{}'.format(k)] = sheet2['C{}'.format(j)].value
            sheet4['D{}'.format(k)] = sheet2['D{}'.format(j)].value
            sheet4['E{}'.format(k)] = sheet2['E{}'.format(j)].value

            if abs(float(sheet4['B{}'.format(k)].value) - b2) <= 0.1:
                sheet4['B{}'.format(k)] = b2
            if abs(float(sheet4['C{}'.format(k)].value) - c2) <= 0.1:
                sheet4['C{}'.format(k)] = c2
            k+=1
    row4 = sheet4.max_row
    # print(row4_1)
    # print(row4)
    # if type(sheet4['B{}'.format(row4_1+1)].value) is not int:
    #     print(type(sheet4['B{}'.format(row4_1+1)].value))
    # if type(sheet4['B{}'.format(row4_1+1)].value) is float:
    # if abs(float(sheet4['B{}'.format(row4_1+1)].value)-b2)<=0.1:
    #     sheet4['B{}'.format(row4_1+1)] = b2
    # if type(sheet4['C{}'.format(row4)].value) is float:
    # if abs(float(sheet4['C{}'.format(row4)].value)-c2)<=0.1:
    #     sheet4['C{}'.format(row4)] = c2
    y = [b2,c2]
    for p in range(row4_1+1,row4+1):
        y.append(sheet4['B{}'.format(p)].value)
        y.append(sheet4['C{}'.format(p)].value)
    # print(y)
    x_sorted = sorted(y)
    x_sorted_kind_2 = []
    for i in x_sorted:
        if (i not in x_sorted_kind_2)and(i>=b2)and(i<=c2):
            x_sorted_kind_2.append(i)
    # print(x_sorted_kind_2)
    row3 = sheet3.max_row
    sheet3['A{}'.format(row3 + 1)] = a2
    sheet3['B{}'.format(row3 + 1)] = b2
    sheet3['C{}'.format(row3 + 1)] = x_sorted_kind_2[1]
    for m in range(2,len(x_sorted_kind_2)):
        row3 = sheet3.max_row
        sheet3['A{}'.format(row3 + 1)] = sheet3['A{}'.format(row3)].value
        sheet3['B{}'.format(row3 + 1)] = sheet3['C{}'.format(row3)].value
        sheet3['C{}'.format(row3 + 1)] = x_sorted_kind_2[m]
        # print(row3)
    # sheet4.delete_rows(2, row4)此行代码无用，加上会报错
wb.save(r"D:\03 python代码\03 区间细化\区间信息导入模版-庆阳市 - 1.xlsx")